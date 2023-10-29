# -*- coding: utf-8 -*-

class step00_create_excel():
    def __init__(self):
        pass

    def run(self, msg):
        # 1. Автомобили Kia в кредит - расчитать кредит на покупку https://www.kia.ru/buy/calc/
        # 2. Расчет стоимости ТО - https://www.kia.ru/service/calculator_to/
        # 3. Программа Trade-In - https://www.kia.ru/buy/trade-in/
        import configparser
        import pathlib
        import os
        import openpyxl
        config = configparser.ConfigParser()
        config.read('config.txt')
        knowledge_dir = config["DEFAULT"]["knowledge_dir"]
        if knowledge_dir is None: knowledge_dir = "./knowledge"
        pathlib.Path(knowledge_dir).mkdir(parents=True, exist_ok=True)
        filepath_01 = os.path.join(knowledge_dir, "01_calc_credit.xlsx")
        wb_01 = openpyxl.Workbook()
        for wn_01 in ["Модели","Двигатель и КПП","Комплектация","Условия кредита","Результаты расчета"]:
            wb_01.create_sheet(wn_01)
            ws_01 = wb_01.worksheets[0]
            wd_01 = {1: "N_", 2: "Модельный ряд", 3: "Комплектация", 4: "Цена покупки (руб)"}
            for col_idx_01 in range(1, 10):
                ws_01.cell(col_idx_01, 1).value = wd_01[col_idx_01] if col_idx_01 in wd_01 else "_"
            for col_idx_01 in range(1, 10):
                col_01 = col_idx_01
                for row_01 in range(2, 10):
                    ws_01.cell(col_01, row_01).value = "?"
        wb_01.save(filepath_01)
        filepath_02 = os.path.join(knowledge_dir, "02_calc_tech.xlsx")
        wb_02 = openpyxl.Workbook()
        for wn_02 in ["Введите Госномер", "Введите VIN-номер"]:
            wb_02.create_sheet(wn_02)
            ws_02 = wb_02.worksheets[0]
            wd_02 = {1: "N_", 2: "Модельный ряд", 3: "Комплектация", 4: "Цена покупки (руб)"}
            for col_idx_02 in range(1, 10):
                ws_02.cell(col_idx_02, 1).value = wd_02[col_idx_02] if col_idx_02 in wd_02 else "_"
            for col_idx_02 in range(1, 10):
                col_02 = col_idx_02
                for row_02 in range(2, 10):
                    ws_02.cell(col_02, row_02).value = "?"
        wb_02.save(filepath_02)
        filepath_03 = os.path.join(knowledge_dir, "03_calc_tradein.xlsx")
        wb_03 = openpyxl.Workbook()
        for wn_03 in ["Ваш автомобиль", "Характеристики вашего авто", "Результаты оценки"]:
            wb_03.create_sheet(wn_03)
            ws_03 = wb_03.worksheets[0]
            wd_03 = {1: "N_", 2: "Модельный ряд", 3: "Комплектация", 4: "Цена покупки (руб)"}
            for col_idx_03 in range(1, 10):
                ws_03.cell(col_idx_03, 1).value = wd_03[col_idx_03] if col_idx_03 in wd_03 else "_"
            for col_idx_03 in range(1, 10):
                col_03 = col_idx_03
                for row_03 in range(2, 10):
                    ws_03.cell(col_03, row_03).value = "?"
        wb_03.save(filepath_03)
        print(msg, " ... OK")

if __name__ == '__main__':
    step00_create_excel().run("")
